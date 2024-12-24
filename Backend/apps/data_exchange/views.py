from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from datetime import datetime
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from apps.applications.models import Application
from apps.applications.serializers import ApplicationSerializer

logger = logging.getLogger(__name__)

class ExcelImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        try:
            logger.info("Starting Excel import process")
            
            if not request.FILES:
                return Response({'error': 'No file uploaded'}, status=400)

            excel_file = request.FILES.get('file')
            if not excel_file:
                return Response({'error': 'No file named "file" in request'}, status=400)
            
            # Read Excel file
            df = pd.read_excel(excel_file)
            logger.info(f"Read DataFrame with columns: {df.columns.tolist()}")
            logger.info(f"First row of data: {df.iloc[0].to_dict() if not df.empty else 'Empty DataFrame'}")

            # Process each row
            applications = []
            errors = []
            for index, row in df.iterrows():
                try:
                    application_data = {
                        'company_name': str(row['company_name']).strip(),
                        'job_title': str(row['job_title']).strip(),
                        'position': str(row['job_title']).strip(),  # Using job_title as position
                        'status': str(row['status']).strip(),
                        'date_applied': pd.to_datetime(row['date_applied']).date(),
                        'job_description': str(row.get('job_description', '')).strip(),
                    }
                    
                    logger.info(f"Processing row {index + 1}")
                    logger.info(f"Data: {application_data}")
                    
                    # Add request to serializer context
                    serializer = ApplicationSerializer(
                        data=application_data,
                        context={'request': request}
                    )
                    
                    if serializer.is_valid():
                        application = serializer.save()
                        applications.append(application)
                        logger.info(f"Successfully created application {application.id}")
                    else:
                        error_detail = {
                            'row': index + 1,
                            'errors': serializer.errors,
                            'data': application_data
                        }
                        errors.append(error_detail)
                        logger.error(f"Row {index + 1} validation failed: {serializer.errors}")

                except Exception as row_error:
                    errors.append({
                        'row': index + 1,
                        'error': str(row_error),
                        'data': row.to_dict()
                    })
                    logger.error(f"Error processing row {index + 1}: {str(row_error)}")
                    continue

            response_data = {
                'message': f'Processed {len(df)} rows',
                'imported_count': len(applications),
                'success_count': len(applications),
                'error_count': len(errors),
            }
            
            if errors:
                response_data['errors'] = errors

            logger.info(f"Import completed. Created {len(applications)} applications")
            return Response(response_data)

        except Exception as e:
            logger.error(f"Error importing Excel file: {str(e)}")
            return Response({'error': str(e)}, status=500)

class ExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get user's applications
            applications = Application.objects.filter(user=request.user)
            
            # Convert to DataFrame
            data = []
            for app in applications:
                data.append({
                    'Company': app.company_name,
                    'Job Title': app.job_title,
                    'Position': app.position,
                    'Status': app.status,
                    'Date Applied': app.date_applied.strftime('%Y-%m-%d') if app.date_applied else None,
                    'Job Description': app.job_description or '',
                    'Created': app.created_at.strftime('%Y-%m-%d %H:%M:%S') if app.created_at else None,
                    'Updated': app.updated_at.strftime('%Y-%m-%d %H:%M:%S') if app.updated_at else None
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            excel_file = BytesIO()
            
            # Create Excel writer
            with pd.ExcelWriter(
                excel_file,
                engine='openpyxl',
                datetime_format='YYYY-MM-DD',
                date_format='YYYY-MM-DD'
            ) as writer:
                df.to_excel(writer, index=False, sheet_name='Applications')
                
                # Get the worksheet
                worksheet = writer.sheets['Applications']
                
                # Style configuration
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                centered_alignment = Alignment(horizontal='center')
                
                # Set minimum column widths
                min_widths = {
                    'Company': 15,
                    'Job Title': 20,
                    'Position': 20,
                    'Status': 12,
                    'Date Applied': 12,
                    'Job Description': 40,
                    'Created': 18,
                    'Updated': 18
                }
                
                # Style headers and adjust column widths
                for idx, col in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = centered_alignment
                    
                    # Get column letter
                    column_letter = get_column_letter(idx)
                    
                    # Calculate max content width
                    max_length = max(
                        min_widths[col],
                        len(str(col)),
                        df[col].astype(str).apply(len).max()
                    )
                    
                    # Add some padding
                    adjusted_width = max_length + 2
                    
                    # Set column width
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Center align date columns
                date_columns = ['Date Applied', 'Created', 'Updated']
                for col in date_columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    for cell in worksheet[col_letter][1:]:
                        cell.alignment = centered_alignment
            
            excel_file.seek(0)
            
            # Prepare response
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'applications_export_{timestamp}.xlsx'
            
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response

        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return Response({'error': str(e)}, status=500) 